# V3/services/batch_builder.py
# Batch PDF builder — clean port of Scripts/make_print_stack.py.
#
# Scans CLEAN folder, groups PDFs by AWB, builds numbered batch PDFs
# with barcode cover pages into data/OUT/.
#
# Key features:
#   - Each PDF opened once (page count collected in scan, not again in builder)
#   - Atomic batch file write (tmp -> rename) — no partial files on crash
#   - Append-mode Excel sequence log — history preserved across runs
#   - Tier batching (ENABLE_TIER_BATCHING=True) — separate PDFs per confidence tier
#   - --estimate-batches CLI flag — reports expected count without building
#   - Centralized audit via V3.audit.tracker.write_batch_event()
#   - MD5 dedup when copying to PENDING_PRINT
#
# All paths and tuning values come from V3.config.

from __future__ import annotations

import csv
import hashlib
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

from V3 import config
from V3.audit.logger import audit_event
from V3.audit.tracker import write_batch_event

try:
    import pymupdf as fitz  # PyMuPDF ≥ 1.24 preferred namespace
except ImportError:
    try:
        import fitz  # type: ignore[no-redef]
        fitz.open  # verify it's real PyMuPDF, not the stub package
    except (ImportError, AttributeError) as exc:
        raise RuntimeError(
            "PyMuPDF import failed. Install PyMuPDF and remove conflicting 'fitz' package."
        ) from exc

from openpyxl import load_workbook, Workbook

# ── Config aliases ────────────────────────────────────────────────────────────
CLEAN_DIR            = config.CLEAN_DIR
OUT_DIR              = config.OUT_DIR
PENDING_PRINT_DIR    = config.PENDING_PRINT_DIR
SEQUENCE_XLSX        = config.SEQUENCE_XLSX
MAX_PAGES_PER_BATCH  = config.MAX_PAGES_PER_BATCH
COVER_PAGE_SIZE      = config.COVER_PAGE_SIZE
PRINT_STACK_BASENAME = config.PRINT_STACK_BASENAME
ENABLE_TIER_BATCHING = config.ENABLE_TIER_BATCHING
STAGE_CACHE_CSV      = config.STAGE_CACHE_CSV

# Matches: 123456789012.pdf  OR  123456789012_2.pdf  OR  123456789012_3.pdf
_AWB_FROM_FILENAME = re.compile(r"^(\d{12})(?:_\d+)?\.pdf$", re.IGNORECASE)

# Tier definitions based on AWB_Detection_Type prefix in stage_cache.csv
_TIER_HIGH_PREFIXES   = ("FILENAME", "TEXTLAYER-EXACT", "TEXT-LAYER")
_TIER_MEDIUM_PREFIXES = ("OCR-EXACT",)


def _require_reportlab() -> bool:
    try:
        import reportlab  # noqa: F401
        return True
    except Exception:
        return False


# =========================
# STAGE CACHE — tier lookup
# =========================

def _load_stage_cache_tiers() -> dict[str, str]:
    """Return ``{awb: tier_label}`` dict from stage_cache.csv."""
    tiers: dict[str, str] = {}
    try:
        if not STAGE_CACHE_CSV.exists():
            return tiers
        with open(STAGE_CACHE_CSV, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                awb = (row.get("AWB_Detected") or "").strip()
                method = (row.get("AWB_Detection_Type") or "").strip().upper()
                if not awb:
                    continue
                if any(method.startswith(p) for p in _TIER_HIGH_PREFIXES):
                    tiers[awb] = "High"
                elif any(method.startswith(p) for p in _TIER_MEDIUM_PREFIXES):
                    tiers[awb] = "Medium"
                else:
                    tiers[awb] = "Low"
    except Exception:
        pass
    return tiers


# =========================
# CLEAN FOLDER SCAN
# =========================

def scan_clean_folder() -> list[dict]:
    """Scan CLEAN_DIR for PDFs grouped by AWB.

    Returns list of dicts:
      ``{awb, pdf_paths, page_counts, mtime_first}``
    sorted by mtime of the first (oldest) file in the AWB group.
    """
    groups: dict[str, list[Path]] = {}
    if not CLEAN_DIR.is_dir():
        return []

    for fn in CLEAN_DIR.iterdir():
        m = _AWB_FROM_FILENAME.match(fn.name)
        if not m:
            continue
        awb = m.group(1)
        groups.setdefault(awb, []).append(fn)

    # Cache stat results to avoid repeated os.stat calls per file
    _mtime_cache: dict[Path, float] = {}
    for awb in groups:
        for p in groups[awb]:
            if p not in _mtime_cache:
                try:
                    _mtime_cache[p] = p.stat().st_mtime
                except Exception:
                    _mtime_cache[p] = 0.0
        groups[awb].sort(key=lambda p: _mtime_cache.get(p, 0.0))

    sorted_groups = sorted(groups.items(), key=lambda kv: _mtime_cache.get(kv[1][0], 0.0))

    result = []
    for awb, pdf_paths in sorted_groups:
        page_counts = []
        valid_paths = []
        for pdf_path in pdf_paths:
            doc = None
            try:
                doc = fitz.open(str(pdf_path))
                pc = doc.page_count
                page_counts.append(pc)
                valid_paths.append(pdf_path)
            except Exception as e:
                print(f"  [WARN] Could not open {pdf_path.name}: {e}")
            finally:
                if doc is not None:
                    try:
                        doc.close()
                    except Exception:
                        pass
        if valid_paths:
            result.append({
                "awb":         awb,
                "pdf_paths":   valid_paths,
                "page_counts": page_counts,
                "mtime_first": _mtime_cache.get(valid_paths[0], 0.0),
            })

    return result


# =========================
# BATCH PLAN
# =========================

def precompute_batch_plan(resolved: list[dict]) -> dict[int, int]:
    """Assign batch numbers to *resolved* entries in-place.

    Returns ``{batch_no: total_pages}`` dict.
    """
    batch_no = 1
    pages_in_current_batch = 0

    for r in resolved:
        sp = r["total_pages"]
        if sp > MAX_PAGES_PER_BATCH:
            print(f"  [WARN] {' | '.join(r.get('pdf_names', ['?']))} has {sp} pages "
                  f"> MAX_PAGES_PER_BATCH={MAX_PAGES_PER_BATCH}; placing alone in its own batch")
            if pages_in_current_batch > 0:
                batch_no += 1
                pages_in_current_batch = 0
        if pages_in_current_batch > 0 and (pages_in_current_batch + sp > MAX_PAGES_PER_BATCH):
            batch_no += 1
            pages_in_current_batch = 0
        r["batch_no"] = batch_no
        r["_batch_start_page"] = pages_in_current_batch + 1
        pages_in_current_batch += sp

    batch_totals: dict[int, int] = {}
    for r in resolved:
        batch_totals[r["batch_no"]] = batch_totals.get(r["batch_no"], 0) + r["total_pages"]

    for r in resolved:
        r["_pages_in_batch"] = batch_totals[r["batch_no"]]
        r["_cover_page_in_batch"] = r["_batch_start_page"]

    return batch_totals


# =========================
# BARCODE COVER PAGE
# =========================

def make_barcode_cover_pdf_bytes(
    awb: str,
    seq: int,
    batch_no: int,
    page_in_batch: int,
    pages_in_batch: int,
    doc_count: int,
    total_inv_pages: int,
    tier: str | None = None,
) -> bytes:
    """Generate a single-page PDF with barcode + metadata as raw bytes."""
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.graphics.barcode import code128

    buf = BytesIO()
    pagesize = letter if COVER_PAGE_SIZE == "LETTER" else A4
    c = canvas.Canvas(buf, pagesize=pagesize)
    w, h = pagesize

    c.setFont("Helvetica-Bold", 18)
    c.drawString(60, h - 80, f"SEQ: {seq}")
    c.setFont("Helvetica-Bold", 22)
    c.drawString(60, h - 120, f"AWB: {awb}")
    c.setFont("Helvetica-Bold", 14)
    c.drawString(60, h - 150, f"BATCH: {batch_no:03d}")
    c.drawString(60, h - 170, f"PAGE: {page_in_batch} of {pages_in_batch}")
    c.setFont("Helvetica", 12)
    c.drawString(60, h - 195, f"Documents: {doc_count}")
    if tier:
        c.drawString(60, h - 215, f"Detection Tier: {tier}")

    barcode = code128.Code128(awb, barHeight=60, barWidth=1.2)
    barcode.drawOn(c, 60, h - 290)

    c.setFont("Helvetica", 10)
    c.drawString(60, 40, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    c.setFont("Helvetica", 7)
    c.drawRightString(w - 60, 24, f"p: {total_inv_pages}")
    c.showPage()
    c.save()
    return buf.getvalue()


# =========================
# EXCEL SEQUENCE LOG (append mode)
# =========================

_SEQ_HEADERS = [
    "Seq", "AWB", "PDF Files", "Timestamp", "DocCount",
    "InvoicePages", "TotalPages", "Batch", "Tier",
]


def write_excel_sequence(resolved: list[dict]) -> None:
    """Append rows to the sequence Excel log (creates if missing)."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if SEQUENCE_XLSX.exists():
        try:
            wb = load_workbook(SEQUENCE_XLSX)
            ws = wb.active
            # Ensure Tier column exists (migration)
            if ws.max_row >= 1 and ws.cell(1, len(_SEQ_HEADERS)).value != "Tier":
                ws.cell(1, len(_SEQ_HEADERS)).value = "Tier"
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sequence"
            ws.append(_SEQ_HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sequence"
        ws.append(_SEQ_HEADERS)

    for r in resolved:
        ws.append([
            r["seq"],
            r["awb"],
            " | ".join(r["pdf_names"]),
            r["timestamp"],
            r["doc_count"],
            r["inv_pages"],
            r["total_pages"],
            r["batch_no"],
            r.get("tier", ""),
        ])
    _tmp = SEQUENCE_XLSX.with_suffix(".xlsx.tmp")
    wb.save(_tmp)
    wb.close()
    os.replace(_tmp, SEQUENCE_XLSX)


# =========================
# ESTIMATE BATCH COUNT (no build)
# =========================

def _estimate_single_stream(groups: list[dict]) -> int:
    """Simulate batch packing for one ordered list of groups; return batch count."""
    batch_no = 1
    pages_in_current = 0
    for g in groups:
        awb_pages = 1 + sum(g["page_counts"])   # 1 cover + invoice pages
        if pages_in_current > 0 and pages_in_current + awb_pages > MAX_PAGES_PER_BATCH:
            batch_no += 1
            pages_in_current = 0
        pages_in_current += awb_pages
    return batch_no


def estimate_batch_count() -> int:
    """Return the expected number of batch PDFs if batching ran now.

    Mirrors the actual build path: when ENABLE_TIER_BATCHING is True each tier
    is a separate series, so the estimate sums across tiers.
    """
    groups = scan_clean_folder()
    if not groups:
        return 0

    if ENABLE_TIER_BATCHING:
        tier_map = _load_stage_cache_tiers()
        total = 0
        for tier_label in ("High", "Medium", "Low"):
            tier_groups = [g for g in groups if tier_map.get(g["awb"], "Low") == tier_label]
            if tier_groups:
                total += _estimate_single_stream(tier_groups)
        return total

    return _estimate_single_stream(groups)


# =========================
# BATCH BUILDER
# =========================

def _batch_filename(tier_label: str | None, batch_no: int) -> str:
    if ENABLE_TIER_BATCHING and tier_label:
        safe = tier_label[0].upper()   # H, M, L
        return f"{PRINT_STACK_BASENAME}_T{safe}_{batch_no:03d}.pdf"
    return f"{PRINT_STACK_BASENAME}_{batch_no:03d}.pdf"


def _save_batch_pdf_atomic(doc, batch_no: int, tier_label: str | None = None) -> Path:
    """Save batch PDF atomically: write to .tmp then os.replace to final name."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    final_name = _batch_filename(tier_label, batch_no)
    out_path   = OUT_DIR / final_name
    tmp_path   = OUT_DIR / (final_name + ".tmp")
    doc.save(str(tmp_path))
    doc.close()
    try:
        os.replace(str(tmp_path), str(out_path))   # atomic on Windows + macOS
    except Exception:
        try:
            tmp_path.unlink()
        except Exception:
            pass
        raise
    return out_path


def build_print_stacks_batched(
    resolved: list[dict],
) -> list[tuple[Path, int, int, str]]:
    """Build batch PDFs from *resolved* entries.

    Returns list of ``(output_path, awb_count, page_count, tier)`` tuples.
    """
    precompute_batch_plan(resolved)

    outputs: list[tuple[Path, int, int, str]] = []
    current_batch_no = None
    batch_doc = None
    batch_pages = 0
    current_tier = None
    batch_awb_count = 0

    for r in resolved:
        target_batch = r["batch_no"]
        tier = r.get("tier", "")

        if current_batch_no is None:
            current_batch_no = target_batch
            batch_doc = fitz.open()
            batch_pages = 0
            current_tier = tier
            batch_awb_count = 0

        if target_batch != current_batch_no:
            out = _save_batch_pdf_atomic(batch_doc, current_batch_no, current_tier)
            outputs.append((out, batch_awb_count, batch_pages, current_tier))
            current_batch_no = target_batch
            batch_doc = fitz.open()
            batch_pages = 0
            current_tier = tier
            batch_awb_count = 0

        # One barcode cover per AWB
        cover_bytes = make_barcode_cover_pdf_bytes(
            awb=r["awb"],
            seq=r["seq"],
            batch_no=r["batch_no"],
            page_in_batch=r["_cover_page_in_batch"],
            pages_in_batch=r["_pages_in_batch"],
            doc_count=r["doc_count"],
            total_inv_pages=r["inv_pages"],
            tier=r.get("tier"),
        )
        cover_doc = fitz.open("pdf", cover_bytes)
        batch_doc.insert_pdf(cover_doc)
        cover_doc.close()
        batch_pages += 1

        for pdf_path in r["pdf_paths"]:
            try:
                inv_doc = fitz.open(str(pdf_path))
                batch_doc.insert_pdf(inv_doc)
                inv_doc.close()
            except Exception as e:
                print(f"  [WARN] Could not insert {pdf_path.name}: {e}")

        batch_pages += r["inv_pages"]
        batch_awb_count += 1

    if batch_doc is not None and batch_pages > 0:
        out = _save_batch_pdf_atomic(batch_doc, current_batch_no, current_tier)
        outputs.append((out, batch_awb_count, batch_pages, current_tier))

    return outputs


def _build_print_stacks_tiered(
    resolved: list[dict],
) -> list[tuple[Path, int, int, str]]:
    """When ENABLE_TIER_BATCHING=True, build separate batch series per tier."""
    outputs_all: list[tuple[Path, int, int, str]] = []
    for tier_label in ("High", "Medium", "Low"):
        tier_resolved = [r for r in resolved if r.get("tier") == tier_label]
        if not tier_resolved:
            continue
        # Re-assign seq numbers within each tier
        for i, r in enumerate(tier_resolved, start=1):
            r["seq"] = i
        outputs_all.extend(build_print_stacks_batched(tier_resolved))
    return outputs_all


# =========================
# SEND BATCHES TO PENDING_PRINT (with MD5 dedup)
# =========================

def _file_md5(path: Path) -> str:
    """Compute MD5 hex digest of a file."""
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def copy_batches_to_pending_print(
    outputs: list[tuple[Path, int, int, str]],
) -> dict:
    """Copy batch PDFs to PENDING_PRINT, skipping exact MD5 duplicates.

    Returns ``{copied, failed, skipped_dup, expected}``.
    """
    PENDING_PRINT_DIR.mkdir(parents=True, exist_ok=True)
    copied = 0
    failed = 0
    skipped_dup = 0

    for src, awb_count, page_count, tier in outputs:
        dst = PENDING_PRINT_DIR / src.name

        # MD5 dedup: skip copy if an identical file is already there
        if dst.exists():
            try:
                if _file_md5(src) == _file_md5(dst):
                    print(f"  [PENDING_PRINT] Skip (MD5 match): {src.name}")
                    skipped_dup += 1
                    continue
            except Exception:
                pass
            # Different content — append version suffix
            stem = src.stem
            suffix = src.suffix
            k = 2
            while True:
                candidate = PENDING_PRINT_DIR / f"{stem}_v{k}{suffix}"
                if not candidate.exists():
                    dst = candidate
                    break
                k += 1

        try:
            shutil.copy2(src, dst)
            copied += 1
            print(f"  [PENDING_PRINT] Copied: {src.name} -> {dst.name}")
            audit_event(
                "BATCH",
                action="copy_to_pending_print",
                source=str(src),
                destination=str(dst),
                status="OK",
            )
            try:
                _bn_match = re.search(r"_(\d+)\.pdf$", src.name, re.IGNORECASE)
                _batch_no = int(_bn_match.group(1)) if _bn_match else None
                write_batch_event(
                    event_type="BATCH_BUILT",
                    batch_number=_batch_no,
                    filename=dst.name,
                    awb_count=awb_count,
                    page_count=page_count,
                    detection_tier_label=tier or "Mixed",
                    output_path=str(dst),
                )
            except Exception:
                pass
        except Exception as e:
            print(f"  [WARN] Could not copy {src.name} to PENDING_PRINT: {e}")
            failed += 1
            audit_event(
                "BATCH",
                action="copy_to_pending_print",
                source=str(src),
                destination=str(dst),
                status="ERROR",
                reason=str(e),
            )
    print(
        f"PENDING_PRINT updated: {copied} file(s) copied."
        + (f" ({failed} failed)" if failed else "")
        + (f" ({skipped_dup} skipped-dup)" if skipped_dup else "")
    )
    return {
        "copied": copied,
        "failed": failed,
        "skipped_dup": skipped_dup,
        "expected": len(outputs),
    }


# =========================
# DELETE CLEAN SOURCES
# =========================

def delete_clean_sources(resolved: list[dict]) -> None:
    """Delete source PDFs in CLEAN after successful batching."""
    deleted = 0
    failed = 0
    for r in resolved:
        for pdf_path in r["pdf_paths"]:
            try:
                if pdf_path.exists():
                    pdf_path.unlink()
                    deleted += 1
                    print(f"  [CLEAN] Deleted: {pdf_path.name}")
            except Exception as e:
                print(f"  [WARN] Could not delete {pdf_path.name}: {e}")
                failed += 1
    print(
        f"Cleaned {deleted} file(s) from CLEAN."
        + (f" ({failed} failed)" if failed else "")
    )


# =========================
# MAIN
# =========================

def main() -> None:
    run_start = time.perf_counter()
    config.ensure_dirs()

    if not _require_reportlab():
        print("ERROR: reportlab not installed. Run: pip install reportlab")
        return

    scanned = scan_clean_folder()

    if not scanned:
        print("No PDFs found in CLEAN folder. Nothing to batch.")
        return

    total_files = sum(len(g["pdf_paths"]) for g in scanned)
    print(f"Found {len(scanned)} AWB(s) in CLEAN ({total_files} file(s) total)")

    # Always load tier_map for audit logging; ENABLE_TIER_BATCHING only controls
    # whether separate per-tier PDFs are built — not whether tier data is recorded.
    tier_map = _load_stage_cache_tiers()

    resolved: list[dict] = []
    seq = 1
    for g in scanned:
        awb        = g["awb"]
        pdf_paths  = g["pdf_paths"]
        page_counts = g["page_counts"]
        inv_pages  = sum(page_counts)
        tier       = tier_map.get(awb, "Low")

        resolved.append({
            "seq":         seq,
            "awb":         awb,
            "timestamp":   datetime.now().isoformat(timespec="seconds"),
            "pdf_paths":   pdf_paths,
            "pdf_names":   [p.name for p in pdf_paths],
            "doc_count":   len(pdf_paths),
            "inv_pages":   inv_pages,
            "total_pages": 1 + inv_pages,   # 1 cover + invoice pages
            "batch_no":    "",
            "tier":        tier,
        })
        seq += 1

    if not resolved:
        print("No readable PDFs found in CLEAN. Nothing to batch.")
        return

    print(f"Building batches for {len(resolved)} AWB(s)...")
    if ENABLE_TIER_BATCHING:
        tier_counts = {
            t: sum(1 for r in resolved if r.get("tier") == t)
            for t in ("High", "Medium", "Low")
        }
        print(
            f"  Tier breakdown: High={tier_counts['High']} "
            f"Medium={tier_counts['Medium']} Low={tier_counts['Low']}"
        )
        outputs = _build_print_stacks_tiered(resolved)
    else:
        outputs = build_print_stacks_batched(resolved)

    write_excel_sequence(resolved)
    copy_result = copy_batches_to_pending_print(outputs)
    # Delete CLEAN sources only if every output was either copied or skipped as dup
    # (i.e., zero failures and the sum of copied+skipped_dup equals expected).
    if (
        copy_result["failed"] == 0
        and (copy_result["copied"] + copy_result["skipped_dup"]) == copy_result["expected"]
    ):
        delete_clean_sources(resolved)
    else:
        print(
            "[SAFETY] Skipping CLEAN source deletion because not all batch files were copied "
            f"to PENDING_PRINT (copied={copy_result['copied']} "
            f"skipped_dup={copy_result['skipped_dup']} "
            f"failed={copy_result['failed']} expected={copy_result['expected']})."
        )

    total_ms = round((time.perf_counter() - run_start) * 1000, 1)
    print("\nDONE")
    print(f"Excel sequence: {SEQUENCE_XLSX}")
    for out_path, awb_count, page_count, tier in outputs:
        print(f"  Batch PDF: {out_path}")

    audit_event(
        "BATCH",
        action="build_print_stacks",
        status="DONE",
        awb_count=len(resolved),
        output_count=len(outputs),
        outputs=[str(t[0]) for t in outputs],
        sequence_xlsx=str(SEQUENCE_XLSX),
        total_active_ms=total_ms,
    )


if __name__ == "__main__":
    if "--estimate-batches" in sys.argv:
        # Lightweight mode: print expected batch count and exit (used by UI auto mode)
        config.ensure_dirs()
        print(estimate_batch_count())
        sys.exit(0)
    main()
