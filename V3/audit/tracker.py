# V3/audit/tracker.py
# UNIFIED audit tracker — replaces both pipeline_tracker.py AND centralized_audit.py.
#
# Single 4-sheet Excel workbook:
#   HotfolderV2  – one row per AWB detection event
#   EDM          – one row per EDM duplicate-check event
#   BatchTIFF    – one row per batch-build or TIFF-convert event
#   Dashboard    – programmatically computed summary (rewritten on every write)
#
# Concurrent write safety: lock file pattern using os.O_CREAT|O_EXCL (atomic
# on both Windows NTFS and macOS APFS).  No new pip dependencies.
#
# Also exposes the pipeline_tracker compat API that the hotfolder calls:
#   record_hotfolder_start(), record_hotfolder_end(), record_hotfolder_needs_review()
# These delegate to write_hotfolder_event() internally.

from __future__ import annotations

import os
import shutil
import sys
import time
import zipfile
from datetime import datetime, date
from pathlib import Path

from V3 import config

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────
_AUDIT_XLSX  = config.AUDIT_XLSX_PATH
_LOCK_FILE   = config.DATA_DIR / "pipeline_audit.lock"
_LOCK_TIMEOUT = 30   # seconds to wait before declaring timeout (raised from 15 for parallel EDM workers)
_LOCK_STALE   = 60   # seconds before an unclaimed lock is considered stale

SHEET_HOT   = "HotfolderV2"
SHEET_EDM   = "EDM"
SHEET_BATCH = "BatchTIFF"
SHEET_DASH  = "Dashboard"

# Column definitions per sheet (name, width)
_HOT_COLS = [
    ("Timestamp",        18), ("EmployeeID",     12), ("AWB",            15),
    ("OriginalFilename", 30), ("Route",          22), ("DetectionMethod", 22),
    ("DetectionTier",    14), ("HotfolderSecs",  14), ("RescueSecs",     14),
    ("Result",           14), ("Notes",          45),
]
_EDM_COLS = [
    ("Timestamp",    18), ("EmployeeID",    12), ("AWB",            15),
    ("Filename",     28), ("EDMResult",     16), ("DupPageCount",   14),
    ("TotalPages",   12), ("DupPct",        10), ("EDMSecs",        12),
    ("CompareMethod",18), ("Notes",         40),
]
_BATCH_COLS = [
    ("Timestamp",    18), ("EmployeeID",    12), ("EventType",      18),
    ("BatchNumber",  14), ("Filename",      28), ("AWBCount",       10),
    ("PageCount",    12), ("DetectionTier", 14), ("Notes",          40),
]

_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_HDR_FONT  = Font(color="FFFFFF", bold=True)
_DASH_FILL = PatternFill("solid", fgColor="2E4057")
_DASH_FONT = Font(color="FFFFFF", bold=True)


# ── Lock helpers (cross-platform) ─────────────────────────────────────────────

def _acquire_lock() -> int:
    """Block until lock acquired or timeout.  Returns fd (int) of the lock file."""
    # Ensure the directory exists before trying to create the lock file inside it.
    # This guards the edge case where a service subprocess starts before ensure_dirs() runs.
    try:
        config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    deadline = time.time() + _LOCK_TIMEOUT
    while True:
        try:
            fd = os.open(str(_LOCK_FILE), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode())
            return fd
        except FileExistsError:
            # Check for stale lock
            try:
                age = time.time() - _LOCK_FILE.stat().st_mtime
                if age > _LOCK_STALE:
                    try:
                        _LOCK_FILE.unlink()
                    except Exception:
                        pass
                    continue
            except Exception:
                pass
        if time.time() > deadline:
            raise TimeoutError(
                f"[tracker] Could not acquire lock after {_LOCK_TIMEOUT}s"
            )
        time.sleep(0.1)


def _release_lock(fd: int) -> None:
    try:
        os.close(fd)
    except Exception:
        pass
    try:
        _LOCK_FILE.unlink()
    except Exception:
        pass


# ── Workbook bootstrap ────────────────────────────────────────────────────────

def _style_header(ws, cols):
    ws.append([c[0] for c in cols])
    for col_idx, (_, width) in enumerate(cols, start=1):
        cell = ws.cell(1, col_idx)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _create_workbook():
    """Create a brand-new pipeline_audit.xlsx with all 4 sheets."""
    wb = Workbook()
    ws_hot = wb.active
    ws_hot.title = SHEET_HOT
    _style_header(ws_hot, _HOT_COLS)

    ws_edm = wb.create_sheet(SHEET_EDM)
    _style_header(ws_edm, _EDM_COLS)

    ws_batch = wb.create_sheet(SHEET_BATCH)
    _style_header(ws_batch, _BATCH_COLS)

    ws_dash = wb.create_sheet(SHEET_DASH)
    _init_dashboard(ws_dash)

    return wb


def _is_recoverable_workbook_error(exc: Exception) -> bool:
    """Return True for errors that mean the xlsx itself cannot be loaded."""
    return isinstance(exc, (zipfile.BadZipFile, InvalidFileException, KeyError, EOFError))


def _quarantine_corrupt_workbook(exc: Exception) -> Path | None:
    """Move a corrupt audit workbook aside so a clean one can be created.

    This runs only while the audit lock is held.  The bad file is preserved for
    later inspection, and any failure here is allowed to bubble up to the
    caller so audit recovery never silently deletes evidence.
    """
    if not _AUDIT_XLSX.exists():
        return None

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    quarantine_dir = config.DATA_DIR / "corrupt_audit_workbooks"
    quarantine_dir.mkdir(parents=True, exist_ok=True)
    quarantine_path = quarantine_dir / f"{_AUDIT_XLSX.stem}.corrupt.{stamp}{_AUDIT_XLSX.suffix}"

    try:
        _AUDIT_XLSX.replace(quarantine_path)
    except Exception:
        shutil.copy2(_AUDIT_XLSX, quarantine_path)
        _AUDIT_XLSX.unlink()

    try:
        sys.stderr.write(
            f"[tracker] Recovered corrupt audit workbook: {exc}. "
            f"Moved bad copy to {quarantine_path}\n"
        )
    except Exception:
        pass
    return quarantine_path


def _open_or_create():
    """Load existing workbook or create new one.  Ensures all 4 sheets exist."""
    config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    if _AUDIT_XLSX.exists():
        try:
            wb = load_workbook(_AUDIT_XLSX)
        except Exception as e:
            if not _is_recoverable_workbook_error(e):
                raise
            _quarantine_corrupt_workbook(e)
            wb = _create_workbook()
        # Ensure all sheets exist (first-run migration)
        for sheet_name, cols in [
            (SHEET_HOT, _HOT_COLS),
            (SHEET_EDM, _EDM_COLS),
            (SHEET_BATCH, _BATCH_COLS),
        ]:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                _style_header(ws, cols)
        if SHEET_DASH not in wb.sheetnames:
            ws_dash = wb.create_sheet(SHEET_DASH)
            _init_dashboard(ws_dash)
    else:
        wb = _create_workbook()
    return wb


def _save_workbook_atomic(wb) -> None:
    """Save the workbook via same-directory temp file, then atomically replace.

    A direct openpyxl save can leave a partial zip if the process is interrupted
    mid-write.  Saving to a temp sibling first keeps the active audit workbook
    untouched until a complete replacement is ready.
    """
    config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp_path = _AUDIT_XLSX.with_name(f".{_AUDIT_XLSX.name}.{os.getpid()}.tmp")
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, _AUDIT_XLSX)
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass


# ── Dashboard ─────────────────────────────────────────────────────────────────

def _init_dashboard(ws):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22


def _rebuild_dashboard(wb):
    """Recompute today's and all-time stats and overwrite the dashboard."""
    ws_dash = wb[SHEET_DASH]
    ws_dash.delete_rows(1, ws_dash.max_row or 1)

    today = date.today().isoformat()

    # ── Count from HotfolderV2 (today + all-time in one pass) ────────────────
    hot_total = hot_complete = hot_review = hot_failed = 0
    at_hot_total = at_hot_complete = at_hot_review = at_hot_failed = 0
    hot_secs_list = []
    tier_counts = {"High": 0, "Medium": 0, "Low": 0}
    ws_hot = wb[SHEET_HOT]
    for row in ws_hot.iter_rows(min_row=2, values_only=True):
        result = str(row[9] or "").upper()
        if result == "IN-PROGRESS":
            continue
        # All-time
        at_hot_total += 1
        if result == "COMPLETE":
            at_hot_complete += 1
        elif result == "NEEDS_REVIEW":
            at_hot_review += 1
        elif result == "FAILED":
            at_hot_failed += 1
        # Today only
        ts = str(row[0] or "")
        if not ts.startswith(today):
            continue
        hot_total += 1
        if result == "COMPLETE":
            hot_complete += 1
        elif result == "NEEDS_REVIEW":
            hot_review += 1
        elif result == "FAILED":
            hot_failed += 1
        tier = str(row[6] or "")
        if tier in tier_counts:
            tier_counts[tier] += 1
        secs = row[7]
        if secs and isinstance(secs, (int, float)):
            hot_secs_list.append(float(secs))

    avg_secs = f"{sum(hot_secs_list)/len(hot_secs_list):.1f}s" if hot_secs_list else "N/A"

    # ── Count from EDM (today + all-time) ────────────────────────────────────
    edm_clean = edm_rejected = edm_partial = edm_unchecked = 0
    at_edm_clean = at_edm_rejected = at_edm_partial = at_edm_unchecked = 0
    ws_edm = wb[SHEET_EDM]
    for row in ws_edm.iter_rows(min_row=2, values_only=True):
        result = str(row[4] or "").upper()
        if result == "CLEAN":
            at_edm_clean += 1
        elif result == "REJECTED":
            at_edm_rejected += 1
        elif result == "PARTIAL-CLEAN":
            at_edm_partial += 1
        elif result == "CLEAN-UNCHECKED":
            at_edm_unchecked += 1
        ts = str(row[0] or "")
        if not ts.startswith(today):
            continue
        if result == "CLEAN":
            edm_clean += 1
        elif result == "REJECTED":
            edm_rejected += 1
        elif result == "PARTIAL-CLEAN":
            edm_partial += 1
        elif result == "CLEAN-UNCHECKED":
            edm_unchecked += 1

    edm_total    = edm_clean + edm_rejected + edm_partial + edm_unchecked
    at_edm_total = at_edm_clean + at_edm_rejected + at_edm_partial + at_edm_unchecked
    edm_clean_rate = (
        f"{(edm_clean + edm_partial) / edm_total * 100:.0f}%"
        if edm_total else "N/A"
    )
    at_edm_clean_rate = (
        f"{(at_edm_clean + at_edm_partial) / at_edm_total * 100:.0f}%"
        if at_edm_total else "N/A"
    )

    # ── Count from BatchTIFF (today + all-time) ───────────────────────────────
    batches_built = tiffs_converted = tiffs_failed = 0
    at_batches_built = at_tiffs_converted = at_tiffs_failed = 0
    ws_batch = wb[SHEET_BATCH]
    for row in ws_batch.iter_rows(min_row=2, values_only=True):
        etype = str(row[2] or "").upper()
        if etype == "BATCH_BUILT":
            at_batches_built += 1
        elif etype == "TIFF_CONVERTED":
            at_tiffs_converted += 1
        elif etype == "TIFF_FAILED":
            at_tiffs_failed += 1
        ts = str(row[0] or "")
        if not ts.startswith(today):
            continue
        if etype == "BATCH_BUILT":
            batches_built += 1
        elif etype == "TIFF_CONVERTED":
            tiffs_converted += 1
        elif etype == "TIFF_FAILED":
            tiffs_failed += 1

    last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # rows = list of (label, today_val, alltime_val, note)
    def _write_section(ws, title, rows):
        title_row = (ws.max_row or 0) + 1
        ws.cell(title_row, 1).value = title
        ws.cell(title_row, 1).font  = _DASH_FONT
        ws.cell(title_row, 1).fill  = _DASH_FILL
        ws.cell(title_row, 2).fill  = _DASH_FILL
        ws.cell(title_row, 3).fill  = _DASH_FILL
        ws.cell(title_row, 4).fill  = _DASH_FILL
        for label, today_val, at_val, note in rows:
            r = (ws.max_row or 0) + 1
            ws.cell(r, 1).value = label
            ws.cell(r, 2).value = today_val
            ws.cell(r, 3).value = at_val
            if note:
                ws.cell(r, 4).value = note

    # Header row with column labels
    hdr_r = 1
    ws_dash.cell(hdr_r, 1).value = f"  DASHBOARD  —  {today}  (updated {last_updated})"
    ws_dash.cell(hdr_r, 1).font  = _DASH_FONT
    ws_dash.cell(hdr_r, 1).fill  = _DASH_FILL
    ws_dash.cell(hdr_r, 2).value = "TODAY"
    ws_dash.cell(hdr_r, 2).font  = _DASH_FONT
    ws_dash.cell(hdr_r, 2).fill  = _DASH_FILL
    ws_dash.cell(hdr_r, 3).value = "ALL TIME"
    ws_dash.cell(hdr_r, 3).font  = _DASH_FONT
    ws_dash.cell(hdr_r, 3).fill  = _DASH_FILL
    ws_dash.cell(hdr_r, 4).fill  = _DASH_FILL

    _write_section(ws_dash, "  AWB HOTFOLDER", [
        ("Files Processed",       hot_total,    at_hot_total,    None),
        ("  Complete",            hot_complete, at_hot_complete, None),
        ("  Needs Review",        hot_review,   at_hot_review,   "manual check required" if hot_review else None),
        ("  Failed",              hot_failed,   at_hot_failed,   "check pipeline.log"    if hot_failed else None),
        ("Avg Processing Time",   avg_secs,     "—",             None),
        ("Tier High (Filename/TextLayer)", tier_counts["High"],   "—", None),
        ("Tier Medium (OCR-Exact)",        tier_counts["Medium"], "—", None),
        ("Tier Low (Tolerance/EDM/Other)", tier_counts["Low"],    "—", None),
    ])
    _write_section(ws_dash, "  EDM DUPLICATE CHECK", [
        ("Files Checked",          edm_total,       at_edm_total,       None),
        ("  Clean",                edm_clean,       at_edm_clean,       None),
        ("  Partial-Clean",        edm_partial,     at_edm_partial,     None),
        ("  Rejected",             edm_rejected,    at_edm_rejected,    "duplicates found" if edm_rejected else None),
        ("  Unchecked (no token)", edm_unchecked,   at_edm_unchecked,   None),
        ("Clean Rate",             edm_clean_rate,  at_edm_clean_rate,  None),
    ])
    _write_section(ws_dash, "  BATCH & TIFF", [
        ("Batches Built",          batches_built,   at_batches_built,   None),
        ("TIFFs Converted",        tiffs_converted, at_tiffs_converted, None),
        ("TIFFs Failed",           tiffs_failed,    at_tiffs_failed,    "check logs" if tiffs_failed else None),
    ])

    ws_dash.column_dimensions["A"].width = 38
    ws_dash.column_dimensions["B"].width = 14
    ws_dash.column_dimensions["C"].width = 14
    ws_dash.column_dimensions["D"].width = 30


# ── Detection tier helper ─────────────────────────────────────────────────────

def detection_tier(method: str) -> str:
    """Map a DetectionMethod string to High / Medium / Low.

    Real method names from the pipeline look like:
      "Filename"                      → High
      "TextLayer-Exact-High"          → High
      "TextLayer-Clean-Exact-High"    → High
      "Probe-0-Exact-High"            → High  (rotation probe exact)
      "OCR-Main-PSM6-Exact-High"      → Medium
      "OCR-Strong-PSM11-Exact-Standard" → Medium
      "OCR-Main-PSM6-Tolerance-High"  → Low
      "400-Pattern"                   → Medium
      ""  / None                      → Low
    """
    if not method:
        return "Low"
    m = method.upper()
    # Filename / text-layer exact matches are highest confidence
    if m.startswith("FILENAME") or m.startswith("TEXTLAYER") or m.startswith("TEXT-LAYER"):
        return "High"
    # Rotation probe exact matches are also high confidence
    if m.startswith("PROBE-") and "EXACT" in m:
        return "High"
    # OCR exact matches (any PSM/DPI combo) and 400-pattern are medium
    if "TOLERANCE" not in m and ("EXACT" in m or "-400" in m or "400-PATTERN" in m):
        return "Medium"
    return "Low"


# ── Internal timestamps / employee ────────────────────────────────────────────

def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _employee() -> str:
    """Read employee ID from environment (set by main.py on subprocess launch)."""
    return os.environ.get("PIPELINE_EMPLOYEE_ID", "")


# ── Public write functions ────────────────────────────────────────────────────

def write_hotfolder_event(
    awb: str | None,
    original_filename: str,
    route: str = "",
    detection_method: str = "",
    hotfolder_secs: float | None = None,
    ocr_context_ms: float | None = None,
    result: str = "",
    notes: str = "",
    employee_id: str | None = None,
) -> None:
    """Write a single row to the HotfolderV2 sheet."""
    tier = detection_tier(detection_method)
    emp  = employee_id or _employee()
    row  = [
        _now(), emp, awb, original_filename, route,
        detection_method, tier,
        round(hotfolder_secs, 2) if hotfolder_secs is not None else None,
        round(ocr_context_ms, 1) if ocr_context_ms is not None else None,
        result, notes,
    ]
    _append_row(SHEET_HOT, row)


def write_edm_event(
    awb: str,
    filename: str,
    edm_result: str,
    dup_page_count: int | None,
    total_pages: int | None,
    dup_ratio: float | None,
    edm_secs: float | None,
    compare_method: str,
    notes: str = "",
    employee_id: str | None = None,
) -> None:
    """Write a single row to the EDM sheet."""
    emp = employee_id or _employee()
    dup_pct = round(dup_ratio * 100) if dup_ratio is not None else None
    row = [
        _now(), emp, awb, filename, edm_result,
        dup_page_count, total_pages,
        dup_pct,
        round(edm_secs, 2) if edm_secs is not None else None,
        compare_method, notes,
    ]
    _append_row(SHEET_EDM, row)


def write_batch_event(
    event_type: str,
    batch_number: int | None = None,
    filename: str | None = None,
    awb_count: int | None = None,
    page_count: int | None = None,
    detection_tier_label: str | None = None,
    output_path: str | None = None,  # accepted but no longer written (kept for compat)
    notes: str = "",
    employee_id: str | None = None,
) -> None:
    """Write a single row to the BatchTIFF sheet."""
    emp = employee_id or _employee()
    row = [
        _now(), emp, event_type, batch_number, filename,
        awb_count, page_count, detection_tier_label,
        notes,
    ]
    _append_row(SHEET_BATCH, row)


_last_dashboard_rebuild: float = 0.0
_DASHBOARD_REBUILD_INTERVAL = 300  # seconds — rebuild dashboard at most every 5 minutes


def rebuild_dashboard_now() -> None:
    """Force a full dashboard rebuild.  Called on app startup and manually."""
    global _last_dashboard_rebuild
    # Ensure DATA_DIR exists before any write (defensive: mirrors _acquire_lock guard).
    try:
        config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    fd = None
    try:
        fd = _acquire_lock()
        wb = _open_or_create()
        _rebuild_dashboard(wb)
        _save_workbook_atomic(wb)
        _last_dashboard_rebuild = time.time()
    except Exception as e:
        try:
            sys.stderr.write(f"[tracker] ERROR rebuilding dashboard: {e}\n")
        except Exception:
            pass
    finally:
        if fd is not None:
            _release_lock(fd)


def _append_row(sheet_name: str, row: list) -> None:
    """Acquire lock, append row, save.  Dashboard rebuilt only periodically."""
    global _last_dashboard_rebuild
    fd = None
    try:
        fd = _acquire_lock()
        wb = _open_or_create()
        wb[sheet_name].append(row)
        # Only rebuild dashboard if interval has elapsed
        now = time.time()
        if now - _last_dashboard_rebuild >= _DASHBOARD_REBUILD_INTERVAL:
            _rebuild_dashboard(wb)
            _last_dashboard_rebuild = now
        _save_workbook_atomic(wb)
    except Exception as e:
        # Never break pipeline flow on audit failure
        try:
            sys.stderr.write(f"[tracker] ERROR writing {sheet_name}: {e}\n")
        except Exception:
            pass
    finally:
        if fd is not None:
            _release_lock(fd)


# ── Read-only: snapshot for stats panel ──────────────────────────────────────

def _read_dashboard_stats_once() -> dict | None:
    """Single attempt to read stats.  Returns None on any exception."""
    defaults = {
        "hot_total": 0, "hot_complete": 0, "hot_review": 0, "hot_failed": 0,
        "edm_clean": 0, "edm_rejected": 0, "edm_partial": 0,
        "batches_built": 0, "tiffs_converted": 0,
        "batch_tier_strong": 0, "batch_tier_mix": 0, "batch_tier_weak": 0,
        "avg_secs": "N/A",
    }
    try:
        if not _AUDIT_XLSX.exists():
            return defaults
        wb = load_workbook(_AUDIT_XLSX, read_only=True, data_only=True)
        today = date.today().isoformat()
        stats = defaults.copy()
        hot_secs_list = []

        if SHEET_HOT in wb.sheetnames:
            for row in wb[SHEET_HOT].iter_rows(min_row=2, values_only=True):
                ts = str(row[0] or "")
                if not ts.startswith(today):
                    continue
                result = str(row[9] or "").upper()
                if result == "COMPLETE":
                    stats["hot_total"] += 1
                    stats["hot_complete"] += 1
                elif result == "NEEDS_REVIEW":
                    stats["hot_total"] += 1
                    stats["hot_review"] += 1
                elif result == "FAILED":
                    stats["hot_total"] += 1
                    stats["hot_failed"] += 1
                # IN-PROGRESS rows not counted — open pass, not a completed file
                secs = row[7]
                if secs and isinstance(secs, (int, float)):
                    hot_secs_list.append(float(secs))

        if hot_secs_list:
            stats["avg_secs"] = f"{sum(hot_secs_list)/len(hot_secs_list):.1f}s"

        if SHEET_EDM in wb.sheetnames:
            for row in wb[SHEET_EDM].iter_rows(min_row=2, values_only=True):
                ts = str(row[0] or "")
                if not ts.startswith(today):
                    continue
                result = str(row[4] or "").upper()
                if result == "CLEAN":
                    stats["edm_clean"] += 1
                elif result == "REJECTED":
                    stats["edm_rejected"] += 1
                elif result == "PARTIAL-CLEAN":
                    stats["edm_partial"] += 1

        if SHEET_BATCH in wb.sheetnames:
            for row in wb[SHEET_BATCH].iter_rows(min_row=2, values_only=True):
                ts = str(row[0] or "")
                if not ts.startswith(today):
                    continue
                etype = str(row[2] or "").upper()
                if etype == "BATCH_BUILT":
                    stats["batches_built"] += 1
                    tier = str(row[7] or "").strip().upper()  # detection_tier_label
                    if tier == "HIGH":
                        stats["batch_tier_strong"] += 1
                    elif tier in {"MEDIUM", "MIXED"}:
                        stats["batch_tier_mix"] += 1
                    else:
                        stats["batch_tier_weak"] += 1
                elif etype == "TIFF_CONVERTED":
                    stats["tiffs_converted"] += 1

        wb.close()
        return stats
    except Exception:
        return None


def read_dashboard_stats() -> dict:
    """Return a flat dict of today's key counts for the UI stats panel.

    Reads WITHOUT acquiring a write lock (safe for display polling).
    Retries once after a short pause if the first attempt catches a corrupt
    mid-write snapshot of the xlsx.  Returns safe defaults only as a last resort.
    """
    defaults = {
        "hot_total": 0, "hot_complete": 0, "hot_review": 0, "hot_failed": 0,
        "edm_clean": 0, "edm_rejected": 0, "edm_partial": 0,
        "batches_built": 0, "tiffs_converted": 0,
        "batch_tier_strong": 0, "batch_tier_mix": 0, "batch_tier_weak": 0,
        "avg_secs": "N/A",
    }
    result = _read_dashboard_stats_once()
    if result is None:
        # First attempt failed — the file was likely mid-write.  Wait briefly
        # for the writer to finish, then try once more before giving up.
        time.sleep(0.15)
        result = _read_dashboard_stats_once()
    return result if result is not None else defaults


def read_alltime_stats() -> dict:
    """Return all-time pipeline outcome counts (no date filter).

    Only counts rows with definitive results (COMPLETE / NEEDS_REVIEW / FAILED).
    IN-PROGRESS rows are excluded so the total equals resolved files only.
    """
    defaults = {"all_total": 0, "all_complete": 0, "all_review": 0, "all_failed": 0}
    try:
        if not _AUDIT_XLSX.exists():
            return defaults
        wb = load_workbook(_AUDIT_XLSX, read_only=True, data_only=True)
        stats = defaults.copy()
        if SHEET_HOT in wb.sheetnames:
            for row in wb[SHEET_HOT].iter_rows(min_row=2, values_only=True):
                result = str(row[9] or "").upper()
                if result == "COMPLETE":
                    stats["all_total"] += 1
                    stats["all_complete"] += 1
                elif result == "NEEDS_REVIEW":
                    stats["all_total"] += 1
                    stats["all_review"] += 1
                elif result == "FAILED":
                    stats["all_total"] += 1
                    stats["all_failed"] += 1
        wb.close()
        return stats
    except Exception:
        return defaults


# ═════════════════════════════════════════════════════════════════════════════
# PIPELINE TRACKER COMPAT API
# ═════════════════════════════════════════════════════════════════════════════
# The hotfolder and other stages call these functions.  They delegate to
# write_hotfolder_event() so that everything goes into the unified workbook.

def record_hotfolder_start(original_filename: str) -> None:
    """Record that hotfolder processing has begun — JSONL only, no Excel row."""
    pass  # start events stay in pipeline_audit.jsonl via audit_event() in pipeline.py


def record_hotfolder_end(
    original_filename: str,
    awb: str,
    route: str,
    match_method: str,
    hotfolder_secs: float | None = None,
    ocr_context_ms: float | None = None,
    notes: str | None = None,
) -> None:
    """Record successful hotfolder completion for *original_filename*."""
    write_hotfolder_event(
        awb=awb,
        original_filename=original_filename,
        route=route,
        detection_method=match_method,
        hotfolder_secs=hotfolder_secs,
        ocr_context_ms=ocr_context_ms,
        result="COMPLETE",
        notes=notes or "",
    )


def record_hotfolder_needs_review(
    original_filename: str,
    reason: str,
    hotfolder_secs: float | None = None,
    detection_method: str = "No Match",
    route: str = "NEEDS_REVIEW",
) -> None:
    """Record that *original_filename* could not be matched and needs manual review."""
    write_hotfolder_event(
        awb="",
        original_filename=original_filename,
        route=route,
        detection_method=detection_method,
        hotfolder_secs=hotfolder_secs,
        ocr_context_ms=None,
        result="NEEDS_REVIEW",
        notes=reason,
    )


# ── Self-test ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Testing V3/audit/tracker.py...")
    write_hotfolder_event(
        awb="123456789012",
        original_filename="invoice_001.pdf",
        route="PROCESSED",
        detection_method="OCR-Exact",
        hotfolder_secs=4.3,
        ocr_context_ms=1200.0,
        result="COMPLETE",
        notes="Test event",
    )
    write_edm_event(
        awb="123456789012",
        filename="123456789012.pdf",
        edm_result="CLEAN",
        dup_page_count=0,
        total_pages=3,
        dup_ratio=0.0,
        edm_secs=1.2,
        compare_method="hash",
    )
    write_batch_event(
        event_type="BATCH_BUILT",
        batch_number=1,
        awb_count=5,
        page_count=22,
        detection_tier_label="Mixed",
        output_path="data/OUT/PRINT_STACK_BATCH_001.pdf",
    )
    record_hotfolder_start("test_invoice.pdf")
    record_hotfolder_end(
        "test_invoice.pdf", "987654321098", "PROCESSED",
        "Filename", hotfolder_secs=1.2, ocr_context_ms=0.0,
    )
    record_hotfolder_needs_review("unknown_doc.pdf", "No AWB found")
    stats = read_dashboard_stats()
    print("Dashboard stats:", stats)
    print(f"Audit saved to: {_AUDIT_XLSX}")
