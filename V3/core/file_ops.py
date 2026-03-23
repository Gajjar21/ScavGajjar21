# V3/core/file_ops.py
# All file operations: stability checks, safe moves, MD5 dedup moves,
# timestamped logging, Tesseract validation, Excel AWB loading,
# stage-cache CSV writing, and AWB logs Excel writing.
#
# Extracted from Scripts/awb_hotfolder_V2.py (monolith).
# Every function is a direct, complete port — no logic simplified or removed.

from __future__ import annotations

import csv
import hashlib
import os
import re
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from V3 import config

# ── Config aliases ────────────────────────────────────────────────────────────
AWB_LEN = config.AWB_LEN
PROCESSED_DIR = config.PROCESSED_DIR
NEEDS_REVIEW_DIR = config.NEEDS_REVIEW_DIR
AWB_LOGS_PATH = config.AWB_LOGS_PATH
LOG_DIR = config.LOG_DIR
STAGE_CACHE_CSV = config.STAGE_CACHE_CSV


# =============================================================================
# LOGGING
# =============================================================================

_log_fh = None  # module-level file handle — opened once, held for session


def _get_log_fh():
    """Return (and lazily open) the persistent log file handle."""
    global _log_fh
    if _log_fh is None or _log_fh.closed:
        try:
            LOG_DIR.mkdir(parents=True, exist_ok=True)
            _log_fh = open(config.PIPELINE_LOG, "a", encoding="utf-8", buffering=1)  # line-buffered
        except Exception:
            _log_fh = None
    return _log_fh


def log(msg: str) -> None:
    """Write a timestamped log line to stdout and to the pipeline log file."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        fh = _get_log_fh()
        if fh is not None:
            fh.write(line + "\n")
    except Exception:
        pass


# =============================================================================
# TESSERACT VALIDATION
# =============================================================================

def require_tesseract() -> None:
    """Raise ``FileNotFoundError`` if the configured Tesseract binary is missing."""
    if not config.TESSERACT_PATH.exists():
        raise FileNotFoundError(
            f"Tesseract not found at: {config.TESSERACT_PATH}"
        )


# =============================================================================
# FILE STABILITY CHECK
# =============================================================================

def file_is_stable(path: str, checks: int = 2, delay: float = 0.3) -> bool:
    """Return ``True`` when *path* has a stable, non-zero size across
    consecutive checks separated by *delay* seconds.

    Used to avoid processing a file that is still being written/copied.
    """
    last = -1
    for _ in range(checks):
        try:
            size = os.path.getsize(path)
        except OSError:
            return False
        if size == last and size > 0:
            return True
        last = size
        time.sleep(delay)
    return False


# =============================================================================
# SAFE MOVE
# =============================================================================

def safe_move(src: str, dst_dir) -> None:
    """Move *src* into *dst_dir*, appending a timestamp suffix if a file
    with the same name already exists.
    """
    name = os.path.basename(src)
    dst = Path(dst_dir) / name
    if dst.exists():
        base, ext = os.path.splitext(name)
        dst = Path(dst_dir) / f"{base}_{int(time.time())}{ext}"
    shutil.move(src, dst)


# =============================================================================
# MOVE TO PROCESSED (with MD5 dedup)
# =============================================================================

def move_to_processed_renamed(src: str, awb: str) -> str:
    """Move *src* to ``PROCESSED_DIR/<awb>.pdf``.

    If the destination already exists:
    - Compare MD5 checksums.  If identical, remove the source silently
      (duplicate content) and return the existing destination path.
    - Otherwise append an incrementing suffix (``_2``, ``_3``, ...).

    Returns the final destination path as a string.
    """

    def file_md5(path: str) -> str:
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    dst = PROCESSED_DIR / f"{awb}.pdf"
    if dst.exists():
        try:
            if file_md5(src) == file_md5(str(dst)):
                log(
                    f"DUPLICATE CONTENT for {awb} -- removing source, skipping move."
                )
                try:
                    os.remove(src)
                except Exception:
                    pass
                return str(dst)
        except Exception:
            pass
        k = 2
        while True:
            dst = PROCESSED_DIR / f"{awb}_{k}.pdf"
            if not dst.exists():
                break
            k += 1
    shutil.move(src, dst)
    return str(dst)


# =============================================================================
# EXCEL AWB LOADER
# =============================================================================

def extract_12_digit_numbers_from_any_text(s) -> List[str]:
    """Extract all 12-digit numbers from arbitrary text *s*.

    Handles both bare 12-digit runs and digit sequences separated by
    hyphens/spaces that total 12 digits after stripping non-digits.
    """
    if s is None:
        return []
    s = str(s)
    out: Set[str] = set()
    for m in re.finditer(r"\b\d{12}\b", s):
        out.add(m.group(0))
    for m in re.finditer(r"(\d[\d\-\s]{10,30}\d)", s):
        d = re.sub(r"\D", "", m.group(0))
        if len(d) == AWB_LEN:
            out.add(d)
    return list(out)


def load_awb_set_from_excel(xlsx_path) -> Set[str]:
    """Load a set of 12-digit AWB numbers from every cell of every sheet
    in the Excel workbook at *xlsx_path*.

    Raises ``FileNotFoundError`` if the file does not exist.
    """
    from openpyxl import load_workbook

    if not Path(xlsx_path).exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    awbs: Set[str] = set()
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    for n in extract_12_digit_numbers_from_any_text(cell):
                        if len(n) == AWB_LEN and n.isdigit():
                            awbs.add(n)
    finally:
        wb.close()
    return awbs


def build_buckets(
    awb_set: Set[str],
) -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    """Build prefix and suffix lookup dicts for fast Hamming-distance matching.

    Returns ``(by_prefix, by_suffix)`` where keys are the first/last 4 digits
    and values are lists of AWBs sharing that prefix/suffix.
    """
    by_prefix: Dict[str, List[str]] = {}
    by_suffix: Dict[str, List[str]] = {}
    for a in awb_set:
        by_prefix.setdefault(a[:4], []).append(a)
        by_suffix.setdefault(a[-4:], []).append(a)
    return by_prefix, by_suffix


# =============================================================================
# STAGE CACHE CSV WRITER
# =============================================================================

def append_stage_cache_row(
    input_file: str,
    processed_file: str,
    awb: str,
    detection_type: str,
    awb_extraction_secs: float,
) -> None:
    """Append a row to the stage-cache CSV.

    Creates the file with headers if it does not yet exist.
    """
    headers = [
        "Timestamp",
        "InputFileName",
        "ProcessedFileName",
        "AWB_Detected",
        "AWB_Detection_Type",
        "AWB_Extraction_Seconds",
    ]
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        os.path.basename(input_file),
        os.path.basename(processed_file),
        awb,
        detection_type,
        awb_extraction_secs,
    ]
    try:
        STAGE_CACHE_CSV.parent.mkdir(parents=True, exist_ok=True)
        new_file = not STAGE_CACHE_CSV.exists()
        with open(STAGE_CACHE_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(headers)
            w.writerow(row)
    except Exception as e:
        log(f"[STAGE_CACHE] Warning: could not write stage cache row: {e}")


# =============================================================================
# AWB LOGS EXCEL WRITER
# =============================================================================

_AWB_LOGS_HEADERS = ["AWB", "SourceFile", "Timestamp", "MatchMethod", "Status"]

# ── Buffered AWB Logs writer ─────────────────────────────────────────────────
# Rows are accumulated in a CSV sidecar file and flushed to the Excel workbook
# in batches to avoid a full workbook load/save on every single match.

_AWB_LOGS_CSV_BUFFER = config.DATA_DIR / "awb_logs_buffer.csv"
_AWB_LOGS_FLUSH_THRESHOLD = 10  # flush to Excel after this many buffered rows


def append_to_awb_logs_excel(
    awb: str,
    source_file: str,
    match_method: str,
    status: str = "MATCHED",
) -> None:
    """Buffer a row to CSV sidecar; flush to Excel when threshold is reached."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [awb, os.path.basename(source_file), ts, match_method, status]
    try:
        AWB_LOGS_PATH.parent.mkdir(parents=True, exist_ok=True)
        new_file = not _AWB_LOGS_CSV_BUFFER.exists()
        with open(_AWB_LOGS_CSV_BUFFER, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(_AWB_LOGS_HEADERS)
            w.writerow(row)
    except Exception as e:
        log(f"[AWB_LOGS] Warning: could not buffer row: {e}")
        return

    # Check if we should flush
    try:
        line_count = sum(1 for _ in open(_AWB_LOGS_CSV_BUFFER, encoding="utf-8")) - 1  # minus header
        if line_count >= _AWB_LOGS_FLUSH_THRESHOLD:
            flush_awb_logs_buffer()
    except Exception:
        pass


def flush_awb_logs_buffer() -> None:
    """Flush buffered CSV rows into the AWB Logs Excel workbook."""
    from openpyxl import Workbook, load_workbook

    if not _AWB_LOGS_CSV_BUFFER.exists():
        return
    # Read buffered rows
    rows_to_write: list[list[str]] = []
    try:
        with open(_AWB_LOGS_CSV_BUFFER, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)  # skip header
            for r in reader:
                if r:
                    rows_to_write.append(r)
    except Exception as e:
        log(f"[AWB_LOGS] Warning: could not read buffer: {e}")
        return
    if not rows_to_write:
        try:
            _AWB_LOGS_CSV_BUFFER.unlink(missing_ok=True)
        except Exception:
            pass
        return

    # Write to Excel
    for attempt in range(3):
        try:
            if AWB_LOGS_PATH.exists():
                wb = load_workbook(AWB_LOGS_PATH)
                ws = wb.active
                if ws.max_row == 0 or ws.cell(1, 1).value != "AWB":
                    ws.insert_rows(1)
                    for col, h in enumerate(_AWB_LOGS_HEADERS, start=1):
                        ws.cell(1, col).value = h
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "AWB Logs"
                ws.append(_AWB_LOGS_HEADERS)
            for row in rows_to_write:
                ws.append(row)
            wb.save(AWB_LOGS_PATH)
            wb.close()
            # Clear the buffer file
            try:
                _AWB_LOGS_CSV_BUFFER.unlink(missing_ok=True)
            except Exception:
                pass
            return
        except PermissionError:
            time.sleep(0.4 * (attempt + 1))
        except Exception as e:
            log(f"[AWB_LOGS] Warning: could not flush to AWB_Logs.xlsx: {e}")
            return
    log(f"[AWB_LOGS] AWB_Logs.xlsx locked after retries — {len(rows_to_write)} rows remain buffered.")
